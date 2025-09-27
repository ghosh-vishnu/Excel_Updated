from django.shortcuts import render

# Create your views here.
import os, uuid, threading, random
from pathlib import Path
from django.conf import settings
from django.http import FileResponse, Http404, HttpResponseBadRequest
from rest_framework.decorators import api_view
from rest_framework.response import Response
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


from converter.utils import extractor

# simple in-memory job tracker
JOBS = {}

def _job_dir(job_id: str) -> Path:
    return Path(settings.MEDIA_ROOT) / job_id

def _cleanup_uploaded_files(folder: Path):
    """Clean up uploaded Word files after successful conversion, keeping only the output files."""
    try:
        for file_path in folder.iterdir():
            if file_path.is_file():
                # Keep only the output Excel and CSV files
                if not (file_path.name.endswith('.xlsx') or file_path.name.endswith('.csv')):
                    file_path.unlink()  # Delete the file
                    print(f"Cleaned up: {file_path.name}")
    except Exception as e:
        print(f"Error during cleanup: {e}")

def _delete_job_folder(job_id: str):
    try:
        folder = _job_dir(job_id)
        if folder.exists():
            import shutil
            shutil.rmtree(folder, ignore_errors=True)
            print(f"Deleted job folder: {job_id}")
    except Exception as e:
        print(f"Error deleting job folder {job_id}: {e}")

def _cleanup_old_jobs():
    """Clean up old job folders that are no longer needed."""
    try:
        import time
        current_time = time.time()
        media_root = Path(settings.MEDIA_ROOT)
        
        for job_folder in media_root.iterdir():
            if job_folder.is_dir() and job_folder.name not in JOBS:
                # Check if folder is older than 1 hour (3600 seconds)
                folder_age = current_time - job_folder.stat().st_mtime
                if folder_age > 3600:  # 1 hour
                    import shutil
                    shutil.rmtree(job_folder)
                    print(f"Cleaned up old job folder: {job_folder.name}")
    except Exception as e:
        print(f"Error during old job cleanup: {e}")

# ------------------- Upload API -------------------
@api_view(['POST'])
def upload_files(request):
    # Clean up old job folders before starting new upload
    _cleanup_old_jobs()

    # Support batched uploads: if jobId provided, append files to same job
    incoming_job_id = request.GET.get("jobId")

    if incoming_job_id and incoming_job_id in JOBS:
        job_id = incoming_job_id
        folder = _job_dir(job_id)
        folder.mkdir(parents=True, exist_ok=True)
    else:
        job_id = str(uuid.uuid4())
        folder = _job_dir(job_id)
        folder.mkdir(parents=True, exist_ok=True)

    files = request.FILES.getlist('files')
    if not files:
        return HttpResponseBadRequest("No files uploaded")

    # Debug: Log all files and their attributes
    print(f"DEBUG: Received {len(files)} files for upload")
    for i, f in enumerate(files):
        print(f"DEBUG: File {i}: name={f.name}, webkitRelativePath={getattr(f, 'webkitRelativePath', 'NOT_SET')}")

    # Extract folder name only when creating a NEW job
    if incoming_job_id is None or incoming_job_id not in JOBS:
        folder_name = "Word_Files"  # Default name
        if files and hasattr(files[0], 'name'):
            # Try to get folder name from webkitRelativePath if available
            webkit_path = getattr(files[0], 'webkitRelativePath', None)
            print(f"DEBUG: webkitRelativePath = {webkit_path}")  # Debug log
            
            if webkit_path and '/' in webkit_path:
                folder_name = webkit_path.split('/')[0]
                print(f"DEBUG: Extracted folder name (/) = {folder_name}")  # Debug log
            elif webkit_path and '\\' in webkit_path:
                folder_name = webkit_path.split('\\')[0]
                print(f"DEBUG: Extracted folder name (\\) = {folder_name}")  # Debug log
            else:
                # If no webkitRelativePath, use a sanitized version of the first file's directory
                folder_name = "Word_Files"
                print(f"DEBUG: No webkitRelativePath found, using default = {folder_name}")  # Debug log

        # Sanitize folder name for filename
        import re
        original_folder_name = folder_name
        folder_name = re.sub(r'[^\w\s-]', '', folder_name).strip()
        folder_name = re.sub(r'[-\s]+', '_', folder_name)
        if not folder_name:
            folder_name = "Word_Files"
        
        print(f"DEBUG: Original folder name = {original_folder_name}, Sanitized = {folder_name}")  # Debug log
    else:
        folder_name = JOBS[job_id].get("folder_name", "Word_Files")

    for f in files:
        filename = f.name
        path = folder / filename
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, 'wb+') as dest:
            for chunk in f.chunks():
                dest.write(chunk)

    if incoming_job_id is None or incoming_job_id not in JOBS:
        JOBS[job_id] = {"progress": 0, "done": False, "result": None, "error": None, "folder_name": folder_name, "cancelled": False}
    # For batched appends, keep existing JOBS entry
    return Response({"jobId": job_id})

# ------------------- Worker function -------------------

def _convert_worker(job_id: str):
    try:
        JOBS[job_id]["progress"] = 5
        folder = _job_dir(job_id)

        files_to_process = [f for f in os.listdir(folder) if f.endswith(".docx") and not f.startswith("~$")]
        total_files = len(files_to_process)
        
        if total_files == 0:
            JOBS[job_id]["progress"] = 100
            JOBS[job_id]["done"] = True
            return

        all_data = []
        for i, file in enumerate(files_to_process):
            if JOBS.get(job_id, {}).get("cancelled"):
                JOBS[job_id]["error"] = "cancelled"
                JOBS[job_id]["done"] = True
                _cleanup_uploaded_files(folder)
                return

            path = folder / file
            print(f"Processing {file}...")

            # extract fields
            title = extractor.extract_title(str(path))
            description = extractor.extract_description(str(path))
            toc = extractor.extract_toc(str(path))
            methodology = extractor.extract_methodology_from_faqschema(str(path))
            seo_title = extractor.extract_seo_title(str(path))
            breadcrumb_text = extractor.extract_breadcrumb_text(str(path))
            skucode = extractor.extract_sku_code(str(path))
            urlrp = extractor.extract_sku_url(str(path))
            breadcrumb_schema = extractor.extract_breadcrumb_schema(str(path))
            meta = extractor.extract_meta_description(str(path))
            schema2 = extractor.extract_faq_schema(str(path))
            report = extractor.extract_report_coverage_table_with_style(str(path))

            # ✅ merge description + report
            merged_text = (description or "") + "\n\n" + (report or "")

            # ✅ split into parts
            chunks = extractor.split_into_excel_cells(merged_text)

            row_data = {
                "File": file,
                "Title": title,
            }

            # add merged description parts
            for j, chunk in enumerate(chunks, start=1):
                row_data[f"Description_Part{j}"] = chunk

            # add other fields (without Report, because merged already)
            row_data.update({
                "TOC": toc,
                "Segmentation": "<p>.</p>",
                "Methodology": methodology,
                "Publish_Date": date.today().strftime('%b-%Y').upper(),
                "Image": "",  # Blank image column
                "Currency": "USD",
                "Single Price": 4485,
                "Corporate Price": 6449,
                "skucode": skucode,
                "Total Page": random.randint(150, 200),
                "RID": "",  # Blank RID column after Single Price
                "Date": date.today().strftime("%d-%m-%Y"),
                "Status": "IN",  # Default status
                "Report_Docs": "",  # Report docs column
                "urlNp": urlrp,
                "Meta Description": meta,
                "Meta_Key": ".",  # Meta key with dot
                "Base Year": "2024",
                "history": "2019-2023",
                "Enterprise Price": 8339,
                "SEOTITLE": seo_title,
                "BreadCrumb Text": breadcrumb_text,
                "Schema 1": breadcrumb_schema,
                "Schema 2": schema2,
                "Sub-Category": ""  # Sub-Category column
                # ⚠ Report removed
            })
            
            all_data.append(row_data)
            JOBS[job_id]["progress"] = 5 + int((i + 1) / total_files * 80)

        if JOBS.get(job_id, {}).get("cancelled"):
            JOBS[job_id]["error"] = "cancelled"
            JOBS[job_id]["done"] = True
            _cleanup_uploaded_files(folder)
            return

        df = pd.DataFrame(all_data)

        # enforce column order
        desc_parts = sorted([c for c in df.columns if c.startswith("Description_Part")],
                            key=lambda x: int(x.replace("Description_Part", "")))

        # Separate Description_Part1 and other Description_Parts
        desc_part1 = [c for c in desc_parts if c == "Description_Part1"]
        other_desc_parts = [c for c in desc_parts if c != "Description_Part1"]

        columns_order = ["File", "Title"] + desc_part1 + [
            "TOC", "Segmentation", "Methodology", "Publish_Date", "Image", "Currency",
            "Single Price", "RID", "Corporate Price", "skucode", "Total Page", "Date", "Status", "Report_Docs",
            "urlNp", "Meta Description", "Meta_Key", "Base Year", "history",
            "Enterprise Price", "SEOTITLE", "BreadCrumb Text", "Schema 1", "Schema 2", "Sub-Category"
        ] + other_desc_parts  # Add other Description_Parts at the end

        df = df[[col for col in columns_order if col in df.columns]]

        folder_name = JOBS[job_id].get("folder_name", "Word_Files")
        xlsx_path = folder / f"{folder_name}.xlsx"
        csv_path = folder / f"{folder_name}.csv"

        df.to_excel(xlsx_path, index=False)
        
        # Apply bold formatting to Publish_Date column
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        # Find Publish_Date column index
        publish_date_col = None
        for col_idx, header in enumerate(ws[1], 1):
            if header.value == "Publish_Date":
                publish_date_col = col_idx
                break
        
        # Apply bold formatting to Publish_Date column
        if publish_date_col:
            bold_font = Font(bold=True)
            for row in range(2, ws.max_row + 1):  # Skip header row
                cell = ws.cell(row=row, column=publish_date_col)
                cell.font = bold_font
        
        wb.save(xlsx_path)
        
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")

        JOBS[job_id]["result"] = {"xlsx": str(xlsx_path), "csv": str(csv_path)}
        JOBS[job_id]["progress"] = 100
        JOBS[job_id]["done"] = True

        _cleanup_uploaded_files(folder)

    except Exception as e:
        JOBS[job_id]["error"] = str(e)
        JOBS[job_id]["done"] = True
        _cleanup_uploaded_files(folder)



@api_view(['POST'])
def reset_job(request):
    """Cancel a running job and delete its files. If jobId missing, clear all jobs."""
    job_id = request.GET.get("jobId")
    if job_id:
        if job_id in JOBS:
            JOBS[job_id]["cancelled"] = True
            _delete_job_folder(job_id)
            del JOBS[job_id]
        return Response({"reset": True, "jobId": job_id})
    # No jobId: clear all
    for jid in list(JOBS.keys()):
        JOBS[jid]["cancelled"] = True
        _delete_job_folder(jid)
        del JOBS[jid]
    return Response({"reset": True, "all": True})

# ------------------- Start Conversion -------------------
@api_view(['POST'])
def start_convert(request):
    job_id = request.GET.get("jobId")
    if not job_id or job_id not in JOBS:
        return HttpResponseBadRequest("Invalid jobId")

    t = threading.Thread(target=_convert_worker, args=(job_id,), daemon=True)
    t.start()
    return Response({"started": True})

# ------------------- Progress -------------------
@api_view(['GET'])
def progress(request):
    job_id = request.GET.get("jobId")
    if not job_id or job_id not in JOBS:
        raise Http404("job not found")
    return Response(JOBS[job_id])

# ------------------- Result download -------------------
@api_view(['GET'])
def result_file(request):
    job_id = request.GET.get("jobId")
    if not job_id or job_id not in JOBS:
        raise Http404("job not found")

    fmt = request.GET.get("format", "xlsx").lower()
    path = JOBS[job_id].get("result", {}).get(fmt)
    if not path or not os.path.exists(path):
        raise Http404("result not ready")

    # Get the folder name for the download filename
    folder_name = JOBS[job_id].get("folder_name", "Word_Files")
    print(f"DEBUG: Download request - job_id={job_id}, folder_name={folder_name}, format={fmt}")  # Debug log
    
    if fmt == "csv":
        filename = f"{folder_name}.csv"
        print(f"DEBUG: Downloading CSV file as: {filename}")  # Debug log
        return FileResponse(open(path, "rb"), as_attachment=True, filename=filename, content_type="text/csv")
    else:
        filename = f"{folder_name}.xlsx"
        print(f"DEBUG: Downloading Excel file as: {filename}")  # Debug log
        return FileResponse(open(path, "rb"), as_attachment=True, filename=filename,
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
